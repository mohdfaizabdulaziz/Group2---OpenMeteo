'''
import openpyxl
import os

OUTPUT_FOLDER  = "output"
EXCEL_FILE = os.path.join(OUTPUT_FOLDER, "Air_Quality.xlsx")

HEADERS = [
    "ID",
    "CITY",
    "COUNTRY",
    "PM10",
    "PM2.5",
    "CO",
    "CO2",
    "CREATED_AT"
]

def save_to_excel(data: dict) -> str:
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        next_id = ws.max_row

    else:
        wb = openpyxl.Workbook()
        ws = wb.active

        from openpyxl.styles import Font, Alignment, PatternFill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

        ws.append(HEADERS)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        col_width = [6, 16, 16, 26]
        for i, width in enumerate(col_width, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        next_id = 1

    row = [
        next_id,
        data["city"],
        data["country"],
        data["air_quality"].get("pm10", [None])[0],
        data["air_quality"].get("pm2_5", [None])[0],
        data["air_quality"].get("carbon_monoxide", [None])[0],
        data["air_quality"].get("carbon_dioxide", [None])[0],
        data["air_quality"].get("scraped_at", None),
    ]

    ws.append(row)
    wb.save(EXCEL_FILE)
    return EXCEL_FILE

if __name__ == "__main__":
    from step1_scrap_meteo import fetch_geocode_and_air_quality

    data = fetch_geocode_and_air_quality("Johor Bahru")
    file_path = save_to_excel(data)'''
import openpyxl
import os

OUTPUT_FOLDER  = "output"
EXCEL_FILE = os.path.join(OUTPUT_FOLDER, "Air_Quality.xlsx")
HEADERS = [
    "ID",
    "CITY",
    "COUNTRY",
    "PM10",
    "PM2.5",
    "CO",
    "CO2",
    "CREATED_AT"
]

def flatten_for_excel(data: dict):
    aq = data["air_quality"]
    return {
        "CITY": data.get("city"),
        "COUNTRY": data.get("country"),
        "PM10": aq.get("pm10", [None])[0],
        "PM2.5": aq.get("pm2_5", [None])[0],
        "CO": aq.get("carbon_monoxide", [None])[0],
        "CO2": aq.get("carbon_dioxide", [None])[0],
        "CREATED_AT": aq.get("scraped_at"),
    }

def save_to_excel(data: dict) -> str:
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    # Flatten if needed
    flat = flatten_for_excel(data) if "air_quality" in data else data
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        next_id = ws.max_row
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        from openpyxl.styles import Font, Alignment, PatternFill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        ws.append(HEADERS)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        col_width = [6, 16, 16, 26]
        for i, width in enumerate(col_width, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        next_id = 1
    row = [
        next_id,
        flat["CITY"],
        flat["COUNTRY"],
        flat["PM10"],
        flat["PM2.5"],
        flat["CO"],
        flat["CO2"],
        flat["CREATED_AT"],
    ]
    ws.append(row)
    wb.save(EXCEL_FILE)
    return EXCEL_FILE